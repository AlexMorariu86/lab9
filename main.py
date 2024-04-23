from openpyxl import load_workbook


def concatenate_columns(input_file, output_file, columns_to_concat, cell=None):
    wb = load_workbook(filename=input_file)
    sheet = wb.active
    new_sheet = wb.create_sheet("Concatenated Results")

    # Get the highest row number (excluding header)
    max_row = sheet.max_row
    cell_value = cell.value
    for row_num in range(2, max_row + 1):
        concatenated_value = ""
        for col_index in columns_to_concat:
            cell = sheet.cell(row=row_num, column=col_index + 1)
            if cell.value:
                concatenated_value += (str)cell_value

        # Create a new cell in the new sheet and set its value
        new_cell = new_sheet.cell(row=row_num - 1, column=len(columns_to_concat) + 1)  # Adjust for 0-based indexing
        new_cell.value = concatenated_value

    wb.save(filename=output_file)
    print("Concatenation completed")


# Example usage (replace file paths and column indexes as needed)
input_file = "C:\\Users\\senon\\Desktop\\Book2.xlsx"
output_file = "C:\\Users\\senon\\Desktop\\Book2conct.xlsx"
columns_to_concat = [0, 2]  # Concatenate values from columns A and B

concatenate_columns(input_file, output_file, columns_to_concat)
